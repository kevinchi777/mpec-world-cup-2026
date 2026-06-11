#!/usr/bin/env python3
"""
Export public World Cup picks from the Excel workbook to picks.json.

This version supports the knockout-stage picker by including the participant's
full name in picks.json. The page still displays abbreviated names publicly,
but the Knockout Picks tab can match the full name that the participant types.

Important privacy note:
- picks.json is public on GitHub Pages.
- This version intentionally includes fullName in picks.json.
- Use only if that is acceptable for your small group.

Intended workflow:
1. Microsoft Forms + Power Automate append responses to tblRawResponses in Excel.
2. This script reads the Excel workbook locally from OneDrive.
3. This script deduplicates submissions by full name; later Excel row wins.
4. This script writes public picks.json into the current GitHub repository folder.
5. update_and_push_to_github.bat commits and pushes picks.json to GitHub Pages.

Requirements:
- Python 3
- openpyxl: pip install openpyxl
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Run setup_once.bat, or run: python -m pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Edit this path only if your workbook moves.
# ---------------------------------------------------------------------------
EXCEL_PATH = Path(r"C:\Users\kevin\OneDrive - Mass General Brigham\world_cup_2026_group_stage_challenge_MPEC_vCurrent.xlsx")

# Public JSON written to the folder where this script lives.
OUTPUT_JSON = Path(__file__).resolve().parent / "picks.json"

GROUPS = list("ABCDEFGHIJKL")

EXPECTED_COLUMNS = [
    "Name",
    "GroupA_Raw",
    "GroupB_Raw",
    "GroupC_Raw",
    "GroupD_Raw",
    "GroupE_Raw",
    "GroupF_Raw",
    "GroupG_Raw",
    "GroupH_Raw",
    "GroupI_Raw",
    "GroupJ_Raw",
    "GroupK_Raw",
    "GroupL_Raw",
    "ThirdPlacePicks_Raw",
    "Tiebreaker_GroupStageGoals",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00A0", " ").strip()


def normalize_name_for_dedup(value: str) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).upper()


def public_display_name(full_name: str) -> str:
    """Convert 'Kevin Chi' to 'Kevin C.' for website display."""
    name = clean_text(full_name)
    if not name:
        return ""

    parts = re.split(r"\s+", name)
    if len(parts) == 1:
        return parts[0]

    first = parts[0]
    last = parts[-1]
    initial_match = re.search(r"[A-Za-zÀ-ÖØ-öø-ÿ]", last)
    if not initial_match:
        return first

    return f"{first} {initial_match.group(0).upper()}."


def normalize_team_name(value: str) -> str:
    text = clean_text(value)
    replacements = {
        "Turkiye": "Türkiye",
        "Curacao": "Curaçao",
        "Czechia": "Czech Republic",
        "Korea Republic": "South Korea",
        "USA": "United States",
        "IR Iran": "Iran",
        "Cabo Verde": "Cape Verde",
        "Bosnia and Herzegovina": "Bosnia-Herzegovina",
        "Côte d'Ivoire": "Ivory Coast",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def parse_ranking(value: Any) -> list[str]:
    raw = clean_text(value)
    if not raw:
        return []

    # Forms/Power Automate may separate rankings with commas, semicolons, or line breaks.
    raw = raw.replace("\r", ",").replace("\n", ",").replace("; ", ",").replace(";", ",")
    teams = [normalize_team_name(part) for part in raw.split(",")]
    return [team for team in teams if team]


def parse_third_place_groups(value: Any) -> list[str]:
    raw = clean_text(value)
    if not raw:
        return []

    # Expected strings look like:
    # "Third-place team from Group A"
    matches = re.findall(r"Group\s+([A-L])", raw)
    groups: list[str] = []
    for group in matches:
        if group not in groups:
            groups.append(group)
    return groups


def parse_tiebreaker(value: Any) -> int | float | str:
    if value is None or value == "":
        return ""
    try:
        numeric = float(value)
        if numeric.is_integer():
            return int(numeric)
        return numeric
    except Exception:
        return clean_text(value)


def find_rawresponses_headers(ws) -> tuple[int, dict[str, int]]:
    """
    Find the row containing the RawResponses table headers.
    Returns: (header_row_number, {header_name: 1-based column_index})
    """
    for row in ws.iter_rows(min_row=1, max_row=20):
        values = [clean_text(cell.value) for cell in row]
        if "Name" in values and "GroupA_Raw" in values and "ThirdPlacePicks_Raw" in values:
            header_map = {value: idx + 1 for idx, value in enumerate(values) if value}
            return row[0].row, header_map

    raise RuntimeError(
        "Could not find RawResponses headers. Expected columns like Name, GroupA_Raw, ThirdPlacePicks_Raw."
    )


def get_cell_by_header(ws, row_number: int, header_map: dict[str, int], header: str) -> Any:
    col = header_map.get(header)
    if col is None:
        return None
    return ws.cell(row=row_number, column=col).value


def main() -> None:
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel workbook not found at:\n{EXCEL_PATH}")
        print("\nCheck that the path is correct and that OneDrive has synced the file locally.")
        sys.exit(1)

    print(f"Reading workbook:\n{EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True, read_only=True)

    if "RawResponses" not in wb.sheetnames:
        print("ERROR: Could not find worksheet named RawResponses.")
        print(f"Available sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)

    ws = wb["RawResponses"]
    header_row, header_map = find_rawresponses_headers(ws)

    missing = [col for col in EXPECTED_COLUMNS if col not in header_map]
    if missing:
        print("WARNING: Some expected columns were not found:")
        for col in missing:
            print(f"  - {col}")
        print("The export will continue, but missing fields may be blank.")

    latest_by_name: dict[str, dict[str, Any]] = {}

    for row_number in range(header_row + 1, ws.max_row + 1):
        raw_name = clean_text(get_cell_by_header(ws, row_number, header_map, "Name"))
        if not raw_name:
            continue

        name_key = normalize_name_for_dedup(raw_name)
        if not name_key:
            continue

        display_name = public_display_name(raw_name)

        groups = {
            group: parse_ranking(get_cell_by_header(ws, row_number, header_map, f"Group{group}_Raw"))
            for group in GROUPS
        }

        pick = {
            "fullName": raw_name,
            "name": display_name,
            "displayName": display_name,
            "groups": groups,
            "thirdPlaceGroups": parse_third_place_groups(
                get_cell_by_header(ws, row_number, header_map, "ThirdPlacePicks_Raw")
            ),
            "tiebreakerGoals": parse_tiebreaker(
                get_cell_by_header(ws, row_number, header_map, "Tiebreaker_GroupStageGoals")
            ),
            "_sourceRow": row_number,
        }

        # Later row wins for duplicate names.
        if name_key not in latest_by_name or row_number > latest_by_name[name_key]["_sourceRow"]:
            latest_by_name[name_key] = pick

    picks = []
    for item in latest_by_name.values():
        public_item = {
            "fullName": item["fullName"],
            "name": item["name"],
            "displayName": item["displayName"],
            "groups": item["groups"],
            "thirdPlaceGroups": item["thirdPlaceGroups"],
            "tiebreakerGoals": item["tiebreakerGoals"],
        }
        picks.append(public_item)

    picks.sort(key=lambda item: item["displayName"].lower())

    output = {
        "metadata": {
            "lastUpdated": datetime.now(timezone.utc).isoformat(),
            "participantCount": len(picks),
            "source": "Generated locally from Excel RawResponses",
            "privacy": "Includes fullName in public picks.json so the Knockout Picks tab can load personalized brackets by full-name lookup.",
            "knockoutLookup": {
                "enabled": True,
                "method": "fullName"
            }
        },
        "picks": picks,
    }

    OUTPUT_JSON.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\nSuccess: wrote {OUTPUT_JSON}")
    print(f"Participants exported: {len(picks)}")
    print("\nIMPORTANT:")
    print("  - picks.json now includes fullName for knockout lookup.")
    print("  - The website still displays abbreviated names using displayName.")
    print("  - This is intended for your small internal group.")


if __name__ == "__main__":
    main()
